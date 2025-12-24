"""
Summarize many JSON files into a table (CSV/XLSX) for Google Sheets.

Example:
  python data_check_summary.py --dir /home/user3/data-analysis \
    --glob "easy_examples_confidence_*.json" \
    --out_csv easy_summary.csv \
    --out_xlsx easy_summary.xlsx
"""

import argparse
import csv
import json
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    Workbook = None


PREFIX = "easy_examples_confidence_"


def parse_filename(path: Path):
    """
    Parse:
      easy_examples_confidence_pythia-160m_0.8_1_0.0005.json
    or:
      easy_examples_confidence_0.7_1_5e-05.json

    Returns dict: model, confidence_thr, epoch, learning_rate
    """
    stem = path.stem
    if not stem.startswith(PREFIX):
        return None

    rest = stem[len(PREFIX):]
    parts = rest.split("_")

    # case A) model + conf + epoch + lr  -> 4 parts
    # case B) conf + epoch + lr          -> 3 parts
    # Be robust: decide whether parts[0] is a model or a number(confidence).
    def is_number_like(s: str) -> bool:
        try:
            float(s)
            return True
        except Exception:
            return False

    model = None
    if len(parts) >= 4 and not is_number_like(parts[0]):
        model = parts[0]
        conf = parts[1]
        epoch = parts[2]
        lr = parts[3]
    elif len(parts) >= 3:
        conf = parts[0]
        epoch = parts[1]
        lr = parts[2]
    else:
        return None

    return {
        "model": model or "",
        "confidence_thr": conf,
        "epoch": epoch,
        "learning_rate": lr,
    }


def summarize_json(path: Path):
    """
    Returns:
      json_type, total_examples, conf_min, conf_max, conf_avg, dict_list_counts
    """
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    json_type = type(data).__name__
    total = ""
    conf_min = ""
    conf_max = ""
    conf_avg = ""
    dict_list_counts = ""

    if isinstance(data, list):
        total = len(data)
        confidences = []
        for item in data:
            if isinstance(item, dict) and "confidence" in item and item["confidence"] is not None:
                try:
                    confidences.append(float(item["confidence"]))
                except Exception:
                    pass
        if confidences:
            conf_min = min(confidences)
            conf_max = max(confidences)
            conf_avg = sum(confidences) / len(confidences)

    elif isinstance(data, dict):
        # If dict contains lists, count them; also sum them as "total" if meaningful.
        list_counts = {}
        sum_lists = 0
        for k, v in data.items():
            if isinstance(v, list):
                list_counts[k] = len(v)
                sum_lists += len(v)
        if list_counts:
            dict_list_counts = json.dumps(list_counts, ensure_ascii=False)
            total = sum_lists

    return {
        "json_type": json_type,
        "total_examples": total,
        "conf_min": conf_min,
        "conf_max": conf_max,
        "conf_avg": conf_avg,
        "dict_list_counts": dict_list_counts,
    }


def write_csv(rows, out_csv: Path):
    headers = list(rows[0].keys()) if rows else []
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_xlsx(rows, out_xlsx: Path):
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed. Install it or only use --out_csv.")

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"

    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)

    header_font = Font(bold=True)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    # Light autosize
    for col_idx, h in enumerate(headers, 1):
        max_len = max(len(str(ws.cell(row=i, column=col_idx).value or "")) for i in range(1, ws.max_row + 1))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(10, max_len + 2), 60)

    wb.save(out_xlsx)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", required=True, help="Directory containing JSON files")
    ap.add_argument("--glob", default="easy_examples_confidence_*.json", help="Glob pattern")
    ap.add_argument("--out_csv", default="easy_summary.csv", help="Output CSV path")
    ap.add_argument("--out_xlsx", default="", help="Output XLSX path (optional)")
    args = ap.parse_args()

    base = Path(args.dir)
    files = sorted(base.glob(args.glob))

    rows = []
    for p in files:
        meta = parse_filename(p)
        if meta is None:
            continue

        try:
            stats = summarize_json(p)
            mtime = datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            rows.append({
                "file": str(p),
                "model": meta["model"],
                "confidence_thr": meta["confidence_thr"],
                "epoch": meta["epoch"],
                "learning_rate": meta["learning_rate"],
                "json_type": stats["json_type"],
                "total_examples": stats["total_examples"],
                "conf_min": stats["conf_min"],
                "conf_max": stats["conf_max"],
                "conf_avg": stats["conf_avg"],
                "dict_list_counts": stats["dict_list_counts"],
                "mtime": mtime,
            })
        except Exception as e:
            rows.append({
                "file": str(p),
                "model": meta["model"],
                "confidence_thr": meta["confidence_thr"],
                "epoch": meta["epoch"],
                "learning_rate": meta["learning_rate"],
                "json_type": "ERROR",
                "total_examples": "",
                "conf_min": "",
                "conf_max": "",
                "conf_avg": "",
                "dict_list_counts": f"{type(e).__name__}: {e}",
                "mtime": "",
            })

    if not rows:
        print("No matching files found.")
        return

    out_csv = Path(args.out_csv)
    write_csv(rows, out_csv)
    print(f"[OK] Wrote CSV: {out_csv.resolve()}")

    if args.out_xlsx:
        out_xlsx = Path(args.out_xlsx)
        write_xlsx(rows, out_xlsx)
        print(f"[OK] Wrote XLSX: {out_xlsx.resolve()}")


if __name__ == "__main__":
    main()